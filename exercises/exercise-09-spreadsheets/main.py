"""Exercise 9: Sort employee spreadsheet by experience."""

from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
INPUT_FILE = BASE_DIR / "data" / "employees.xlsx"
OUTPUT_FILE = BASE_DIR / "employees_sorted.xlsx"

EXPECTED_HEADERS = [
    "name",
    "years of experience",
    "job title",
    "date of birth",
]


def read_employees() -> list[tuple[str, float]]:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"Input spreadsheet not found: {INPUT_FILE}"
        )

    workbook = load_workbook(INPUT_FILE, data_only=True)
    worksheet = workbook.active

    headers = [
        worksheet.cell(row=1, column=column).value
        for column in range(1, 5)
    ]

    normalized_headers = [
        str(header).strip().lower()
        if header is not None
        else ""
        for header in headers
    ]

    if normalized_headers != EXPECTED_HEADERS:
        raise ValueError(
            "Unexpected spreadsheet headers.\n"
            f"Expected: {EXPECTED_HEADERS}\n"
            f"Found:    {normalized_headers}"
        )

    employees: list[tuple[str, float]] = []

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        name = row[0]
        experience = row[1]

        if name is None or experience is None:
            continue

        try:
            experience_number = float(experience)
        except (TypeError, ValueError) as error:
            raise ValueError(
                f"Invalid experience value for {name}: "
                f"{experience}"
            ) from error

        employees.append(
            (str(name).strip(), experience_number)
        )

    workbook.close()
    return employees


def write_sorted_employees(
    employees: list[tuple[str, float]],
) -> None:
    sorted_employees = sorted(
        employees,
        key=lambda employee: employee[1],
        reverse=True,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Employees"

    worksheet.append(
        ["name", "years of experience"]
    )

    for name, experience in sorted_employees:
        worksheet.append([name, experience])

    workbook.save(OUTPUT_FILE)


def main() -> None:
    employees = read_employees()
    write_sorted_employees(employees)

    print(
        f"Sorted {len(employees)} employee records."
    )
    print(
        f"Output written to: {OUTPUT_FILE}"
    )


if __name__ == "__main__":
    main()